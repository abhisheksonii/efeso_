import streamlit as st
import pandas as pd
import tempfile
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
from datetime import datetime
from typing import Dict, List, Tuple, Any
import io
import numpy as np
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from io import BytesIO
from openpyxl.cell.cell import MergedCell
import anthropic
import json

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def read_file(uploaded_file: Any) -> pd.ExcelFile:
    """Read different file formats and return as ExcelFile object"""
    file_type = uploaded_file.name.split('.')[-1].lower()
    
    if file_type in ['xlsx', 'xls', 'xlsm', 'xlsb', 'odf', 'ods', 'odt']:
        return pd.ExcelFile(uploaded_file)
    elif file_type == 'csv':
        # Try different encodings
        for encoding in ['utf-8', 'iso-8859-1', 'cp1252']:
            try:
                df = pd.read_csv(uploaded_file, encoding=encoding)
                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                    df.to_excel(writer, index=False)
                buffer.seek(0)
                return pd.ExcelFile(buffer)
            except Exception as e:
                continue
        raise ValueError("Could not read CSV file with any supported encoding")
    else:
        raise ValueError(f"Unsupported file format: {file_type}")

def initialize_claude_client():
    """Initialize Claude Haiku client"""
    try:
        api_key = st.secrets["ANTHROPIC_API_KEY"]
        return anthropic.Client(api_key)
    except Exception as e:
        logger.warning(f"Failed to initialize Claude client: {str(e)}")
        return None



def process_production_data(df: pd.DataFrame, client=None) -> List[Dict[str, Any]]:
    """Process production data with validation and calculations"""
    sections = []
    current_stream = None
    current_section = {}
    
    for idx, row in df.iterrows():
        # Skip empty rows
        if pd.isna(row['Line']) or str(row['Line']).strip() == '':
            continue
            
        process_name = str(row['Asset / Area']).strip()
        line = str(row['Line']).strip()
        
        # New section detection
        if 'Production Stream' in line or 'SRA' in line:
            if current_section:
                sections.append(current_section)
            current_stream = line
            current_section = {
                'stream': current_stream,
                'processes': []
            }
            
        if process_name and process_name not in ['Process', 'Total Tonnes', 'Despatch']:
            try:
                # Validate required fields
                required_fields = ['Nameplate Speed', 'Average Speed (UoM/hr)', 'Production Volume', 
                                 'Number of Machines', 'Standard Manning Days', 'Machine Hours per Week']
                for field in required_fields:
                    if pd.isna(row[field]):
                        logger.warning(f"Missing required field {field} for process {process_name}")
                        continue

                # Calculate reference speed
                ref_speed = float(row['Nameplate Speed']) if not pd.isna(row['Nameplate Speed']) else float(row['Average Speed (UoM/hr)'])
                
                # Calculate value adding hours with validation
                value_adding_hours = 0
                if ref_speed > 0 and float(row['Number of Machines']) > 0:
                    value_adding_hours = (float(row['Production Volume']) / (ref_speed * 60) / float(row['Number of Machines']))
                
                # Calculate direct operators per machine/shift
                dir_op = (float(row['Standard Manning Days']) / float(row['Number of Machines'])) if not pd.isna(row['Standard Manning Days']) else 0
                
                # Calculate machine hours and required manhours
                machine_hours = float(row['Machine Hours per Week']) if not pd.isna(row['Machine Hours per Week']) else 0
                required_manhours = dir_op * machine_hours * float(row['Number of Machines'])
                
                # Calculate saturation with validation
                saturation = min((machine_hours / 168), 1) if machine_hours > 0 else 0
                
                # Calculate OEE dynamically if possible
                try:
                    theoretical_output = ref_speed * machine_hours * 60
                    actual_output = float(row['Production Volume'])
                    oee = actual_output / theoretical_output if theoretical_output > 0 else 0.57
                    oee = min(max(oee, 0), 1)  # Ensure OEE is between 0 and 1
                except:
                    oee = 0.57  # Default value if calculation fails
                
                process_data = {
                    'name': process_name,
                    'machines': int(row['Number of Machines']),
                    'volume': float(row['Production Volume']),
                    'unit': str(row['UoM']),
                    'value_adding_hours': round(value_adding_hours, 2),
                    'ref_speed': round(ref_speed, 2),
                    'oee': round(oee, 2),
                    'actual_hours': round(machine_hours, 2),
                    'saturation': round(saturation, 2),
                    'operators_per_machine': round(dir_op, 2),
                    'required_manhours': round(required_manhours, 2),
                    'actual_manhours': calculate_actual_manhours(row),
                    'org_losses': calculate_organizational_losses(row),
                    'actual_ops': calculate_actual_operators(row)
                }
                
                if current_section:
                    current_section['processes'].append(process_data)
                
            except Exception as e:
                logger.warning(f"Error processing row {idx} - {process_name}: {str(e)}")
                continue
    
    # Add final section if exists
    if current_section:
        sections.append(current_section)
        
    return sections
def calculate_actual_manhours(row: pd.Series) -> float:
    """Calculate actual manhours based on available data"""
    try:
        if not pd.isna(row.get('Actual Manhours')):
            return float(row['Actual Manhours'])
        elif not pd.isna(row.get('Standard Manning Days')) and not pd.isna(row.get('Machine Hours per Week')):
            return float(row['Standard Manning Days']) * float(row['Machine Hours per Week'])
        return 0
    except:
        return 0  
def calculate_organizational_losses(row: pd.Series) -> float:
    """Calculate organizational losses based on available data"""
    try:
        if not pd.isna(row.get('Organizational Losses')):
            return float(row['Organizational Losses'])
        actual_hours = calculate_actual_manhours(row)
        required_hours = float(row['Standard Manning Days']) * float(row['Machine Hours per Week'])
        if required_hours > 0:
            return max(0, (actual_hours - required_hours) / actual_hours)
        return 0
    except:
        return 0

def calculate_actual_operators(row: pd.Series) -> int:
    """Calculate actual number of operators based on available data"""
    try:
        if not pd.isna(row.get('Actual Operators')):
            return int(row['Actual Operators'])
        elif not pd.isna(row.get('Standard Manning Days')):
            return int(round(float(row['Standard Manning Days'])))
        return 0
    except:
        return 0  
def handle_file_processing(file):
    """Process a single file and return the report data"""
    try:
        excel_file = read_file(file)
        client = initialize_claude_client()
        
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            sections = process_production_data(df, client)
            
            if sections:
                capacity_df, labour_df = transform_to_capacity_report(sections)
                report_bytes = export_capacity_report(
                    capacity_df, 
                    labour_df, 
                    file_date=datetime.now().strftime("%Y-%m-%d")
                )
                return capacity_df, labour_df, report_bytes
                
        raise ValueError("No valid data found in the uploaded file")
        
    except Exception as e:
        logger.error(f"Error processing file {file.name}: {str(e)}", exc_info=True)
        raise

def get_claude_analysis(client, df_json: str) -> Dict:
    """Get analysis from Claude Haiku"""
    try:
        prompt = f"""Given this production data in JSON format:
        {df_json}
        
        Please analyze this data and return a JSON object with:
        1. Calculated formulas for each process
        2. Aggregated metrics
        3. Capacity insights
        4. Labour optimization recommendations
        
        Format the response as valid JSON with these keys:
        - formulas
        - metrics
        - capacity_insights
        - labour_recommendations"""
        
        message = client.messages.create(
            model="claude-3-haiku-20240307",
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}]
        )
        
        return json.loads(message.content[0].text)
    except Exception as e:
        logger.warning(f"Claude analysis failed: {str(e)}")
        return {}
    
def transform_to_capacity_report(sections: List[Dict[str, Any]]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Transform processed data into capacity and labour reports with calculations"""
    if not sections:
        raise ValueError("No valid sections to process")
        
    capacity_data = []
    labour_data = []
    
    for section in sections:
        for process in section['processes']:
            # Validate process data
            if not all(key in process for key in ['name', 'machines', 'volume', 'unit', 'value_adding_hours',
                                                'ref_speed', 'oee', 'actual_hours', 'saturation']):
                logger.warning(f"Skipping incomplete process data: {process.get('name', 'Unknown')}")
                continue
                
            # Create capacity row with validated data
            capacity_row = {
                'Process': process['name'],
                '# Mach. Avail.': process['machines'],
                'Production Volume (Weekly)': process['volume'],
                'Meas. Unit': process['unit'],
                'Value Adding Mc Hours/Week/Mc': process['value_adding_hours'],
                'Ref Speed (Meas. Unit per min)': process['ref_speed'],
                'Actual OEE': process['oee'],
                'Actual Mc Hours/Week/Mc': process['actual_hours'],
                'Saturation vs. 168': process['saturation']
            }
            
            # Create labour row with validated data
            labour_row = {
                'Dir op/mach/shift': process['operators_per_machine'],
                'Required Manhours/Week': process['required_manhours'],
                'Actual Manhour/Week': process['actual_manhours'],
                'Org. Losses': process['org_losses'],
                'Actual No of Dir. Ops': process['actual_ops'],
                'Target Prod': calculate_target_productivity(process),
                'Actual Efficiency': calculate_actual_efficiency(process)
            }
            
            capacity_data.append(capacity_row)
            labour_data.append(labour_row)
    
    # Calculate and add totals
    if capacity_data:
        total_row_capacity = calculate_totals(capacity_data)
        total_row_labour = calculate_labour_totals(labour_data)
        
        capacity_data.append(total_row_capacity)
        labour_data.append(total_row_labour)
    
    return pd.DataFrame(capacity_data), pd.DataFrame(labour_data)
def calculate_target_productivity(process: Dict[str, Any]) -> float:
    """Calculate target productivity based on process data"""
    try:
        if process['actual_hours'] > 0 and process['operators_per_machine'] > 0:
            return process['volume'] / (process['actual_hours'] * process['operators_per_machine'])
        return 0
    except:
        return 0

def calculate_actual_efficiency(process: Dict[str, Any]) -> float:
    """Calculate actual efficiency based on process data"""
    try:
        target_prod = calculate_target_productivity(process)
        if target_prod > 0:
            return process['volume'] / (process['actual_manhours'] / process['machines'])
        return 0
    except:
        return 0

def calculate_totals(capacity_data: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Calculate totals for capacity data"""
    total_machines = sum(row['# Mach. Avail.'] for row in capacity_data)
    total_volume = sum(row['Production Volume (Weekly)'] for row in capacity_data)
    avg_oee = np.mean([row['Actual OEE'] for row in capacity_data])
    avg_hours = np.mean([row['Actual Mc Hours/Week/Mc'] for row in capacity_data])
    avg_saturation = np.mean([row['Saturation vs. 168'] for row in capacity_data])
    
    return {
        'Process': 'Total',
        '# Mach. Avail.': total_machines,
        'Production Volume (Weekly)': total_volume,
        'Meas. Unit': '',
        'Value Adding Mc Hours/Week/Mc': '',
        'Ref Speed (Meas. Unit per min)': '',
        'Actual OEE': avg_oee,
        'Actual Mc Hours/Week/Mc': avg_hours,
        'Saturation vs. 168': avg_saturation
    }

def calculate_labour_totals(labour_data: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Calculate totals for labour data"""
    total_required = sum(row['Required Manhours/Week'] for row in labour_data)
    total_actual = sum(row['Actual Manhour/Week'] for row in labour_data)
    total_ops = sum(row['Actual No of Dir. Ops'] for row in labour_data)
    avg_losses = np.mean([row['Org. Losses'] for row in labour_data if row['Org. Losses'] > 0])
    
    return {
        'Dir op/mach/shift': '',
        'Required Manhours/Week': total_required,
        'Actual Manhour/Week': total_actual,
        'Org. Losses': avg_losses,
        'Actual No of Dir. Ops': total_ops,
        'Target Prod': '',
        'Actual Efficiency': ''
    }

def export_capacity_report(capacity_df: pd.DataFrame, labour_df: pd.DataFrame, file_date: str = None) -> bytes:
    """Export capacity report in EFESO template format with proper headers"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Process Data"
    
    # Colors
    header_gray = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    light_blue = PatternFill(start_color='99CCFF', end_color='99CCFF', fill_type='solid')
    
    # Insert header rows
    ws.insert_rows(1, 4)  # Add one more row for sub-headers
    
    # Set EFESO logo
    ws['A1'].value = "EFESO"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    
    # Set date header
    date_text = f"Start Point [{file_date}]" if file_date else "Start Point"
    ws['A3'].value = date_text
    ws['A3'].font = Font(bold=True)
    ws['A3'].fill = header_gray
    ws.merge_cells('A3:P3')
    
    # Set section headers
    ws['A4'].value = "Process"
    ws['B4'].value = "CAPACITY"
    ws.merge_cells('B4:I4')
    ws['J4'].value = "LABOUR"
    ws.merge_cells('J4:P4')
    
    # Write capacity column headers
    capacity_headers = list(capacity_df.columns)
    for idx, header in enumerate(capacity_headers):
        cell = ws.cell(row=5, column=idx + 1)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_gray
        
    # Write labour column headers
    labour_headers = list(labour_df.columns)
    for idx, header in enumerate(labour_headers):
        cell = ws.cell(row=5, column=idx + 10)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_gray
    
    # Write data starting from row 6
    start_row = 6
    for idx, row in capacity_df.iterrows():
        for col_idx, value in enumerate(row):
            cell = ws.cell(row=start_row + idx, column=col_idx + 1)
            cell.value = value
            
    # Write labour data starting from column J (10)
    for idx, row in labour_df.iterrows():
        for col_idx, value in enumerate(row):
            cell = ws.cell(row=start_row + idx, column=col_idx + 10)
            cell.value = value
    
    # Apply styling
    style_excel_sheet(wb, capacity_df, labour_df)
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def style_excel_sheet(wb, capacity_df, labour_df):
    """Apply styling to match the image exactly"""
    ws = wb.active
    
    # Colors
    header_gray = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    light_blue = PatternFill(start_color='99CCFF', end_color='99CCFF', fill_type='solid')
    green_fill = PatternFill(start_color='339966', end_color='339966', fill_type='solid')
    
    # Fonts
    header_font = Font(bold=True, size=11)
    process_font = Font(name='Arial', bold=True, size=24)
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Column widths (matching image)
    column_widths = {
        'A': 30,  # Process
        'B': 8,   # # Mach.
        'C': 12,  # Production Volume
        'D': 10,  # Meas. Unit
        'E': 15,  # Value Adding
        'F': 15,  # Ref Speed
        'G': 10,  # Actual OEE
        'H': 15,  # Actual Mc Hours
        'I': 12,  # Saturation
        'J': 12,  # Dir op
        'K': 15,  # Required Manhours
        'L': 15,  # Actual Manhours
        'M': 10,  # Org. Losses
        'N': 15   # Actual No
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Apply styling to all cells
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, MergedCell):
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Process column styling
                if col == 1:
                    cell.fill = light_blue
                    cell.font = Font(bold=True)
                
                # Saturation column styling
                if col == 9:
                    cell.fill = green_fill
                    cell.font = Font(color='FFFFFF', bold=True)
                    
                # Number formatting
                if col in [7, 9, 13]:  # OEE, Saturation, Org Losses columns
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0%'
                elif col in [2, 3, 10, 11, 12, 14]:  # Numeric columns
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'

def main():
    st.set_page_config(page_title="Production Analysis Dashboard", layout="wide")
    st.title("Production Analysis Dashboard")
    
    uploaded_files = st.file_uploader(
        "Upload production files",
        type=['xlsx', 'xls', 'csv'],
        accept_multiple_files=True
    )
    
    if uploaded_files:
        try:
            # Process files
            for file in uploaded_files:
                excel_file = read_file(file)
                
                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    sections = process_production_data(df)
                    
                    if sections:
                        # Transform data
                        capacity_df, labour_df = transform_to_capacity_report(sections)
                        
                        # Generate report
                        report_bytes = export_capacity_report(capacity_df, labour_df)
                        
                        # Download button
                        st.download_button(
                            label="Download Capacity Report",
                            data=report_bytes,
                            file_name="capacity_report.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                        # Preview
                        with st.expander("Report Preview"):
                            col1, col2 = st.columns(2)
                            with col1:
                                st.write("Capacity Data")
                                st.dataframe(capacity_df)
                            with col2:
                                st.write("Labour Data")
                                st.dataframe(labour_df)
                        break
                        
        except Exception as e:
            st.error(f"Error processing files: {str(e)}")
            logger.error(f"Processing error: {str(e)}", exc_info=True)
            raise

def set_streamlit_config():
    """Configure Streamlit page settings"""
    st.set_page_config(
        page_title="Production Analysis Dashboard",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # Custom CSS to improve layout
    st.markdown("""
        <style>
        .stApp {
            max-width: 1200px;
            margin: 0 auto;
        }
        .stDownloadButton {
            width: 100%;
        }
        </style>
    """, unsafe_allow_html=True)

def display_instructions():
    """Display usage instructions"""
    st.markdown("""
        ### Instructions
        1. Upload your production data file(s) in Excel (.xlsx, .xls) or CSV format
        2. The system will automatically process the data and generate a capacity report
        3. Preview the processed data in the expandable section
        4. Download the formatted report using the download button
        
        **Note**: Files should contain production stream data with proper column headers.
    """)

def handle_file_processing(file):
    """Process a single file and return the report data"""
    try:
        excel_file = read_file(file)
        
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            # Remove any completely empty rows
            df = df.dropna(how='all')
            
            sections = process_production_data(df)
            
            if sections:
                capacity_df, labour_df = transform_to_capacity_report(sections)
                report_bytes = export_capacity_report(
                    capacity_df, 
                    labour_df, 
                    file_date=datetime.now().strftime("%Y-%m-%d")
                )
                return capacity_df, labour_df, report_bytes
                
        raise ValueError("No valid data found in the uploaded file")
        
    except Exception as e:
        logger.error(f"Error processing file {file.name}: {str(e)}", exc_info=True)
        raise

if __name__ == "__main__":
    try:
        # Initialize logging
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.StreamHandler(),
                logging.FileHandler('app.log')
            ]
        )
        
        # Set up Streamlit configuration
        set_streamlit_config()
        
        # Display app title and instructions
        st.title("Production Analysis Dashboard")
        display_instructions()
        
        # File upload section
        uploaded_files = st.file_uploader(
            "Upload production files",
            type=['xlsx', 'xls', 'csv'],
            accept_multiple_files=True
        )
        
        if uploaded_files:
            for file in uploaded_files:
                try:
                    with st.spinner(f'Processing {file.name}...'):
                        capacity_df, labour_df, report_bytes = handle_file_processing(file)
                        
                        # Download button
                        st.download_button(
                            label=f"Download Capacity Report for {file.name}",
                            data=report_bytes,
                            file_name=f"capacity_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                        # Preview section
                        with st.expander(f"Report Preview - {file.name}"):
                            col1, col2 = st.columns(2)
                            with col1:
                                st.write("Capacity Data")
                                st.dataframe(capacity_df)
                            with col2:
                                st.write("Labour Data")
                                st.dataframe(labour_df)
                                
                except Exception as e:
                    st.error(f"Error processing {file.name}: {str(e)}")
                    logger.error(f"Error processing {file.name}: {str(e)}", exc_info=True)
                    continue
                    
    except Exception as e:
        st.error("An unexpected error occurred. Please check the logs for details.")
        logger.error(f"Application error: {str(e)}", exc_info=True)